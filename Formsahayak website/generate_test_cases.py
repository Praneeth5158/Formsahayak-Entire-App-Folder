import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define the dataset of 335 test cases
# Format: (Module, Feature, Test Scenario, Precondition, Test Steps, Test Data, Expected Result, Priority, Severity, Test Type, Automation Candidate, Suites)
tc_list = [
    # A. Authentication Testing (30 Test Cases)
    (
        "Authentication", "Signup", "Verify signup with valid credentials",
        "User is on the Registration screen and not registered",
        "1. Enter unique username\n2. Enter valid email\n3. Enter strong password\n4. Confirm password\n5. Click 'Sign Up'",
        "Username: 'ramesh_g', Email: 'ramesh.g@example.com', Password: 'Password@123', Confirm: 'Password@123'",
        "User account created, redirect to language selection page.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Authentication", "Signup", "Verify signup with already registered email",
        "User is on the Registration screen. Email is already registered in DB.",
        "1. Enter unique username\n2. Enter duplicate email\n3. Enter password\n4. Confirm password\n5. Click 'Sign Up'",
        "Username: 'ramesh_new', Email: 'ramesh.g@example.com', Password: 'Password@123', Confirm: 'Password@123'",
        "Error message displayed: 'Email address already registered.'",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with invalid email format",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email without @ symbol\n3. Enter password\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: 'ramesh.g_example.com', Password: 'Password@123'",
        "Error message: 'Please enter a valid email address.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with password less than 8 characters",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email\n3. Enter short password (6 chars)\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: 'ramesh.g@example.com', Password: 'Pass12'",
        "Error message: 'Password must be at least 8 characters long.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with password lacking uppercase letters",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email\n3. Enter password with no uppercase\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: 'ramesh.g@example.com', Password: 'password@123'",
        "Validation error: 'Password must contain at least one uppercase letter.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with password lacking numbers",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email\n3. Enter password with no numbers\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: 'ramesh.g@example.com', Password: 'Password@'",
        "Validation error: 'Password must contain at least one digit.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with password lacking special characters",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email\n3. Enter password with no special char\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: 'ramesh@example.com', Password: 'Password123'",
        "Validation error: 'Password must contain at least one special character.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with password mismatch in confirmation",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email\n3. Enter password\n4. Enter mismatching confirm password\n5. Click 'Sign Up'",
        "Password: 'Password@123', Confirm: 'Password@124'",
        "Validation error: 'Passwords do not match.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with empty username",
        "User is on the Registration screen",
        "1. Leave username field blank\n2. Fill other registration details\n3. Click 'Sign Up'",
        "Username: '', Email: 'ramesh@example.com', Password: 'Password@123'",
        "Validation error: 'Username is required.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with empty email",
        "User is on the Registration screen",
        "1. Fill username\n2. Leave email blank\n3. Fill password\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: '', Password: 'Password@123'",
        "Validation error: 'Email is required.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify signup with empty password",
        "User is on the Registration screen",
        "1. Fill username and email\n2. Leave password blank\n3. Click 'Sign Up'",
        "Username: 'ramesh', Email: 'ramesh@example.com', Password: ''",
        "Validation error: 'Password is required.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify leading and trailing spaces handling in email",
        "User is on the Registration screen",
        "1. Enter username\n2. Enter email with spaces\n3. Enter password\n4. Click 'Sign Up'",
        "Username: 'ramesh', Email: ' ramesh@example.com ', Password: 'Password@123'",
        "Account created successfully, spaces should be trimmed on backend.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify SQL injection protection in signup username field",
        "User is on the Registration screen",
        "1. Enter SQL payload in username\n2. Enter valid email & password\n3. Click 'Sign Up'",
        "Username: 'john\'; DROP TABLE users;--', Email: 'john@example.com', Password: 'Password@123'",
        "Input sanitized. Registration fails or succeeds as plain text username without executing query.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Signup", "Verify XSS protection in signup username field",
        "User is on the Registration screen",
        "1. Enter HTML/JS tag payload in username\n2. Enter valid email/password\n3. Click 'Sign Up'",
        "Username: '<script>alert(1)</script>', Email: 'john@example.com', Password: 'Password@123'",
        "Script tags should be HTML-encoded and stored as plain text. No script execution in UI.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify login with valid registered email and password",
        "User is registered and is on the Login screen",
        "1. Enter registered email\n2. Enter correct password\n3. Click 'Login'",
        "Email: 'ramesh.g@example.com', Password: 'Password@123'",
        "User successfully logged in, redirect to Dashboard screen with active session.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Authentication", "Login", "Verify login with unregistered email",
        "User is on the Login screen",
        "1. Enter unregistered email\n2. Enter password\n3. Click 'Login'",
        "Email: 'not_exists@example.com', Password: 'Password@123'",
        "Error message: 'Invalid email or password.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify login with incorrect password",
        "User is registered and is on the Login screen",
        "1. Enter registered email\n2. Enter incorrect password\n3. Click 'Login'",
        "Email: 'ramesh.g@example.com', Password: 'WrongPassword'",
        "Error message: 'Invalid email or password.'",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Authentication", "Login", "Verify login with empty email field",
        "User is on the Login screen",
        "1. Leave email blank\n2. Enter password\n3. Click 'Login'",
        "Email: '', Password: 'Password@123'",
        "Validation error: 'Email is required.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify login with empty password field",
        "User is on the Login screen",
        "1. Enter registered email\n2. Leave password blank\n3. Click 'Login'",
        "Email: 'ramesh.g@example.com', Password: ''",
        "Validation error: 'Password is required.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify SQL injection protection in login email field",
        "User is on the Login screen",
        "1. Enter SQL injection payload in email\n2. Click 'Login'",
        "Email: 'admin@example.com\' OR \'1\'=\'1', Password: 'Any'",
        "Error message 'Invalid email or password' displayed; query is not executed.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify login with case-insensitive email check",
        "User is registered on the system with lowercase email",
        "1. Enter email in uppercase\n2. Enter valid password\n3. Click 'Login'",
        "Email: 'RAMESH.G@EXAMPLE.COM', Password: 'Password@123'",
        "Successful login, dashboard loads. Email checks should be case-insensitive.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify login password case-sensitivity",
        "User is registered with password having mixed casing",
        "1. Enter correct email\n2. Enter password with incorrect case\n3. Click 'Login'",
        "Email: 'ramesh.g@example.com', Password: 'password@123'",
        "Error message displayed: 'Invalid email or password.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Logout", "Verify user can log out successfully",
        "User is logged into the application dashboard",
        "1. Tap profile/settings\n2. Tap the 'Logout' button\n3. Confirm logout on confirmation popup",
        "N/A",
        "User session is invalidated, and user is redirected back to the Login screen.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Authentication", "Logout", "Verify back button behavior after logging out",
        "User has logged out and is on the Login screen",
        "1. Press the device/browser back button",
        "N/A",
        "User remains on the Login screen and is not allowed to view the Dashboard.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Regression"]
    ),
    (
        "Authentication", "Session Validation", "Verify direct dashboard access without authentication",
        "User is not logged in",
        "1. Attempt to access dashboard URL / activity directly",
        "URL/Screen: /dashboard",
        "User is redirected to the login screen with a warning: 'Authentication required.'",
        "High", "Critical", "Security", "Yes", ["Smoke", "Regression", "Critical Path"]
    ),
    (
        "Authentication", "Session Validation", "Verify JWT token validation expiration timeout",
        "User is logged in with active session",
        "1. Leave user session idle until token expires (e.g. 24 hours)\n2. Perform any authenticated operation",
        "N/A",
        "Session terminates. User is automatically logged out and redirected to Login screen.",
        "High", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Authentication", "Session Validation", "Verify JWT signature tampering detection",
        "User has an active session token stored locally",
        "1. Intercept API request\n2. Alter JWT token signature payload\n3. Resubmit API request",
        "Modified JWT string",
        "Server rejects request with 401 Unauthorized status and locks session.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify account lock after consecutive failed login attempts",
        "User account exists",
        "1. Enter correct email\n2. Enter incorrect password 5 times in a row",
        "Email: 'ramesh.g@example.com', Password: 'wrong'",
        "Account is locked temporarily for 15 minutes. Error displays: 'Too many login attempts.'",
        "High", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "Authentication", "Session Validation", "Verify concurrent login session limit",
        "User is logged in on Device A",
        "1. Log in with same credentials on Device B\n2. Perform activity on Device A",
        "N/A",
        "Device A session is invalidated or user is prompted to terminate other sessions.",
        "Medium", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Authentication", "Login", "Verify 'Remember Me' persistent session functionality",
        "User is on the Login screen",
        "1. Check 'Remember Me' checkbox\n2. Login with valid credentials\n3. Close and reopen the app",
        "Remember Me checked",
        "User bypasses login screen and is directly shown the dashboard.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),

    # B. Profile Testing (25 Test Cases)
    (
        "Profile", "Edit Profile", "Verify profile details editing with valid data",
        "User is logged in and navigates to the Profile screen",
        "1. Tap 'Edit Profile'\n2. Update full name\n3. Tap 'Save'",
        "New Name: 'Ramesh Goud'",
        "Profile updates successfully; database displays updated name instantly in UI.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Profile", "Edit Profile", "Verify profile editing with invalid email format",
        "User is logged in and on the Edit Profile screen",
        "1. Enter invalid email address\n2. Tap 'Save'",
        "Email: 'ramesh_invalid'",
        "Error message displayed: 'Please enter a valid email address.' Profile not updated.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Edit Profile", "Verify profile editing with empty username/name",
        "User is logged in and on the Edit Profile screen",
        "1. Clear full name field\n2. Tap 'Save'",
        "Name: ''",
        "Validation error: 'Name cannot be blank.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Update Phone", "Verify phone number update with valid 10-digit Indian number",
        "User is logged in and on the Edit Profile screen",
        "1. Enter valid 10-digit number\n2. Tap 'Save/Verify OTP'\n3. Enter valid OTP",
        "Phone: '9876543210', OTP: '123456'",
        "Phone number is successfully updated and OTP verified.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Profile", "Update Phone", "Verify phone number update with less than 10 digits",
        "User is logged in and on the Edit Profile screen",
        "1. Enter 9-digit number\n2. Tap 'Save'",
        "Phone: '987654321'",
        "Validation error: 'Phone number must be exactly 10 digits.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Update Phone", "Verify phone number update with alphanumeric characters",
        "User is logged in and on the Edit Profile screen",
        "1. Enter phone number containing letters\n2. Tap 'Save'",
        "Phone: '98765abc10'",
        "Validation error: 'Phone number can only contain numeric digits.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Update Phone", "Verify phone number update with duplicate registered number",
        "User is logged in and on the Edit Profile screen. Phone number already exists in DB.",
        "1. Enter duplicate phone number\n2. Tap 'Save'",
        "Phone: '9988776655'",
        "Error message: 'Phone number already registered by another user.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Update Language", "Verify language selection change to English",
        "User is logged in and on the Profile Settings screen",
        "1. Select English from language dropdown",
        "Language: English",
        "Application UI text immediately translates to English.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Profile", "Update Language", "Verify language selection change to Telugu",
        "User is logged in and on the Profile Settings screen",
        "1. Select Telugu from language dropdown",
        "Language: Telugu",
        "Application UI text immediately translates to Telugu.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Profile", "Update Language", "Verify language selection change to Hindi",
        "User is logged in and on the Profile Settings screen",
        "1. Select Hindi from language dropdown",
        "Language: Hindi",
        "Application UI text immediately translates to Hindi.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Profile", "Update Language", "Verify language selection change to Tamil",
        "User is logged in and on the Profile Settings screen",
        "1. Select Tamil from language dropdown",
        "Language: Tamil",
        "Application UI text immediately translates to Tamil.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify upload of profile image in valid JPG format",
        "User is logged in and on the Profile screen",
        "1. Click on profile image avatar\n2. Select a valid JPG file (1.2 MB)\n3. Click 'Upload'",
        "File: 'profile.jpg', size 1.2MB",
        "Profile image is updated successfully, rendered in avatar circle.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify upload of profile image in valid PNG format",
        "User is logged in and on the Profile screen",
        "1. Click on profile image avatar\n2. Select a valid PNG file (800 KB)\n3. Click 'Upload'",
        "File: 'profile.png', size 800KB",
        "Profile image is updated successfully, rendered in avatar circle.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image upload exceeding size limit",
        "User is logged in and on the Profile screen. Max size limit is 5MB.",
        "1. Select a JPG file of size 7.5 MB\n2. Click 'Upload'",
        "File: 'large_profile.jpg', size 7.5MB",
        "Error message displayed: 'File size exceeds the maximum limit of 5MB.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image upload with unsupported format (GIF)",
        "User is logged in and on the Profile screen",
        "1. Select a GIF format file\n2. Click 'Upload'",
        "File: 'avatar.gif'",
        "Error message: 'Unsupported file format. Only JPG and PNG are allowed.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image upload with unsupported format (PDF)",
        "User is logged in and on the Profile screen",
        "1. Select a PDF format file\n2. Click 'Upload'",
        "File: 'document.pdf'",
        "Error message: 'Unsupported file format. Only JPG and PNG are allowed.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image upload with unsupported format (SVG)",
        "User is logged in and on the Profile screen",
        "1. Select an SVG format file\n2. Click 'Upload'",
        "File: 'vector.svg'",
        "Error message: 'Unsupported file format. Only JPG and PNG are allowed.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image upload with corrupted image file",
        "User is logged in and on the Profile screen",
        "1. Select a corrupted JPG file\n2. Click 'Upload'",
        "File: 'corrupted.jpg'",
        "Error message: 'Unable to parse the image file. Please upload a valid image.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image upload with empty (0-byte) file",
        "User is logged in and on the Profile screen",
        "1. Select an empty JPG file (0 bytes)\n2. Click 'Upload'",
        "File: 'empty.jpg'",
        "Error message: 'Uploaded file is empty.'",
        "Low", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Edit Profile", "Verify profile details persist after closing and reopening application",
        "User is logged in and has modified profile details",
        "1. Force close the application\n2. Reopen the application\n3. Navigate to Profile screen",
        "N/A",
        "Profile details displayed match the previously updated values.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify profile image cropping and styling configuration",
        "User is logged in and on the Profile screen",
        "1. Upload a rectangular image (not square)\n2. Confirm upload",
        "File: 'rect.jpg'",
        "UI correctly crops/fits the image within the circular avatar frame without distortion.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Profile", "Edit Profile", "Verify offline profile changes error display",
        "User is logged in and cuts internet connection",
        "1. Disconnect internet\n2. Tap 'Edit Profile'\n3. Modify name and tap 'Save'",
        "Name: 'Ramesh Online'",
        "Error toast: 'No internet connection. Please check your network and try again.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Profile", "Edit Profile", "Verify XSS prevention in profile editing name field",
        "User is logged in and on the Edit Profile screen",
        "1. Enter HTML tags with script in name field\n2. Tap 'Save'",
        "Name: '<b onmouseover=alert(1)>John</b>'",
        "Profile name saved, script tags stripped or HTML encoded. No popups trigger.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify image upload with double extension bypass attempt",
        "User is logged in and on the Profile screen",
        "1. Select a file named 'exploit.png.exe' containing shell script\n2. Click 'Upload'",
        "File: 'exploit.png.exe'",
        "Upload fails. Validation checks full filename and actual MIME type.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Profile", "Edit Profile", "Verify canceling of profile changes doesn't persist edits",
        "User is logged in and on the Edit Profile screen",
        "1. Modify profile name\n2. Click the 'Cancel' or 'Back' button\n3. Check profile name",
        "Name: 'Ramesh Cancel'",
        "Changes are discarded, profile displays the old username.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),

    # C. OCR Testing (25 Test Cases)
    (
        "OCR", "OCR Text Extraction", "Verify EasyOCR text extraction on clear English document",
        "User is logged in, has uploaded a clean document image containing printed English text",
        "1. Start OCR extraction on English form\n2. Check output fields",
        "Form: English Bank Opening Form.jpg (High resolution)",
        "EasyOCR parses text accurately (above 95% accuracy) and fields mapped correctly.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify EasyOCR text extraction on clear Telugu document",
        "User is logged in, has uploaded a clean Telugu document",
        "1. Start OCR extraction\n2. Check output fields",
        "Form: Telugu Aadhar Form.jpg (High resolution)",
        "EasyOCR parses Telugu characters accurately and displays translations/overlay.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify EasyOCR text extraction on clear Hindi document",
        "User is logged in, has uploaded a clean Hindi document",
        "1. Start OCR extraction\n2. Check output fields",
        "Form: Hindi PAN Form.jpg (High resolution)",
        "EasyOCR parses Hindi characters accurately and displays correct text in UI.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify EasyOCR text extraction on clear Tamil document",
        "User is logged in, has uploaded a clean Tamil document",
        "1. Start OCR extraction\n2. Check output fields",
        "Form: Tamil Passport Form.jpg (High resolution)",
        "EasyOCR parses Tamil characters accurately and displays correct text in UI.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR warning message for slightly blurry image",
        "User is on the Form Upload page",
        "1. Select a slightly blurry image of a form\n2. Proceed to OCR extraction",
        "File: 'slightly_blurry.png'",
        "OCR triggers, but warning is displayed: 'Quality is low. Field accuracy may be affected.'",
        "Medium", "Major", "Functional", "No", ["Sanity", "Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR failure handling on extremely blurry image",
        "User is on the Form Upload page",
        "1. Select an extremely blurry/unreadable image\n2. Proceed to OCR extraction",
        "File: 'very_blurry.jpg'",
        "Error message displayed: 'Unable to extract text. Please upload a clear photo of the form.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR processing on 90-degree rotated image",
        "User is on the Form Upload page",
        "1. Upload a form image rotated 90 degrees clockwise\n2. Tap 'Process'",
        "File: 'rotated_90.jpg'",
        "Backend automatically rotates the image to upright position or extracts text using orientation tags.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR processing on 180-degree rotated image",
        "User is on the Form Upload page",
        "1. Upload a form image rotated upside down (180 deg)\n2. Tap 'Process'",
        "File: 'rotated_180.jpg'",
        "Backend auto-rotates the page and extracts text with correct bounding box mapping.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR processing on large dimensions high-res image",
        "User is on the Form Upload page",
        "1. Select high-res image (8MB, 5000x5000 px)\n2. Click upload and process",
        "File: 'hi_res.png', size 8MB",
        "System compresses the image locally or handles large payload, executing OCR successfully.",
        "Medium", "Major", "Performance", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR warning for low resolution pixelated image",
        "User is on the Form Upload page",
        "1. Upload a pixelated form image (under 300x300 px)\n2. Tap 'Process'",
        "File: 'pixelated.jpg'",
        "Error or warning message: 'Resolution is too low. Upload an image above 1000px width/height.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR text extraction on image with uneven lighting/shadows",
        "User is on the Form Upload page",
        "1. Upload a form photo taken with heavy shadow across the center\n2. Tap 'Process'",
        "File: 'shadow_form.jpg'",
        "Backend pre-processes image (contrast stretching) to extract text with decent accuracy.",
        "Medium", "Major", "Functional", "No", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR text extraction with complex table cells",
        "User is on the Form Upload page",
        "1. Upload a form having multi-row nested table structures\n2. Tap 'Process'",
        "File: 'table_form.png'",
        "OCR correctly segments cells and returns correct field relationships.",
        "Medium", "Major", "Functional", "No", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR on handwritten form inputs",
        "User is on the Form Upload page",
        "1. Upload a form filled out with handwriting\n2. Tap 'Process'",
        "File: 'handwritten.jpg'",
        "EasyOCR extracts printed structure and guides user; handwritten sections flagged or extracted with alert.",
        "Low", "Minor", "Functional", "No", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR on mixed language (English/Telugu) document",
        "User is on the Form Upload page",
        "1. Upload a form containing both Telugu and English instructions\n2. Tap 'Process'",
        "File: 'bilingual_form.png'",
        "EasyOCR correctly initializes multi-language models (e.g. ['te', 'en']) and extracts both scripts.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR on inverted color document (white text on black background)",
        "User is on the Form Upload page",
        "1. Upload an inverted-color form image\n2. Tap 'Process'",
        "File: 'inverted_form.jpg'",
        "System handles thresholding and successfully reads text elements.",
        "Low", "Minor", "Functional", "No", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify canceling of OCR process mid-operation",
        "User starts OCR processing on a document",
        "1. Click the 'Cancel' or 'Stop' button while processing spinner is visible",
        "N/A",
        "API call is aborted, client loading spinner terminates, resource is cleaned up.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR extraction on cropped/incomplete document page",
        "User is on the Form Upload page",
        "1. Upload a cropped image missing top header\n2. Tap 'Process'",
        "File: 'cropped_form.jpg'",
        "OCR parses available text; guidance system alerts user that form template is incomplete.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR server-side latency timeout handling",
        "User starts OCR processing",
        "1. Simulate high backend delay (e.g., 30s response from Render server)\n2. Check frontend behavior",
        "Network latency set to 35 seconds",
        "App shows error message: 'Server request timed out. Please try processing again.'",
        "High", "Major", "Integration", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify client loading spinner indicator visibility during OCR",
        "User is on the Form page",
        "1. Click 'Extract Text'\n2. Observe screen during execution",
        "N/A",
        "Progress bar or skeleton screen is displayed. All action buttons are disabled during processing.",
        "Medium", "Minor", "UI", "Yes", ["Sanity", "Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify EasyOCR bounding box data accuracy in response",
        "User runs OCR on form",
        "1. Verify returned API JSON response structure",
        "N/A",
        "JSON payload contains keys: 'text', 'coordinates' (x, y, width, height), 'confidence'.",
        "High", "Major", "Integration", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR extraction accuracy on small text font sizes",
        "User uploads document",
        "1. Upload form containing disclaimer section in 6pt font\n2. Tap 'Process'",
        "File: 'small_print.jpg'",
        "Text is extracted if resolution allows; otherwise, user is guided using fallback help notes.",
        "Low", "Minor", "Functional", "No", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR text extraction on wrinkled or folded paper form",
        "User uploads document",
        "1. Upload photo of a wrinkled PAN form\n2. Tap 'Process'",
        "File: 'wrinkled.jpg'",
        "Pre-processing handles shadows from folds. Text is retrieved.",
        "Medium", "Minor", "Functional", "No", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR extraction on a completely blank page",
        "User uploads document",
        "1. Upload a completely blank image\n2. Tap 'Process'",
        "File: 'blank.jpg'",
        "System alerts user: 'No text detected in the uploaded image. Please upload a valid form.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR extraction confidence score filters",
        "User runs OCR",
        "1. Look up confidence score in console/logs for low contrast items",
        "Confidence score < 0.4",
        "Fields with very low confidence are flagged in backend logs or marked for review.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "OCR", "OCR Text Extraction", "Verify OCR processing on non-document photo",
        "User is on the Form page",
        "1. Upload photo of a cat or generic scenery\n2. Tap 'Process'",
        "File: 'cat.jpg'",
        "App identifies image as non-form: 'Invalid document type. Please upload a structured form.'",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),

    # D. Form Upload Testing (25 Test Cases)
    (
        "Form Upload", "Upload Form", "Verify form upload in valid JPG format",
        "User is logged in and on the Form Upload screen",
        "1. Select a valid JPG form (1.5 MB)\n2. Click 'Upload'",
        "File: 'sbi_form.jpg', 1.5MB",
        "Form uploaded successfully. Server saves file, database writes metadata, returns success toast.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Form Upload", "Upload Form", "Verify form upload in valid PNG format",
        "User is logged in and on the Form Upload screen",
        "1. Select a valid PNG form (950 KB)\n2. Click 'Upload'",
        "File: 'pan_form.png', 950KB",
        "Form uploaded successfully. Server saves file, returns status code 201.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify form upload in valid single-page PDF format",
        "User is logged in and on the Form Upload screen",
        "1. Select a single-page PDF file\n2. Click 'Upload'",
        "File: 'aadhar.pdf', single page",
        "Form uploaded successfully. Server converts PDF page to image for EasyOCR processing.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Form Upload", "Upload Form", "Verify form upload in valid multi-page PDF format",
        "User is logged in and on the Form Upload screen",
        "1. Select a 3-page PDF file\n2. Click 'Upload'",
        "File: 'passport_app.pdf', 3 pages",
        "PDF uploaded successfully. Frontend/backend lists each page for step-by-step guidance.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify form upload at the maximum allowed size limit",
        "User is logged in and on the Form Upload screen. Max size limit is 5MB.",
        "1. Select a JPG form file of size exactly 5.0 MB\n2. Click 'Upload'",
        "File: '5mb_form.jpg'",
        "Form uploaded successfully. Processed without error.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify form upload exceeding the maximum size limit",
        "User is logged in. Limit is 5MB.",
        "1. Select a PNG form file of size 5.2 MB\n2. Click 'Upload'",
        "File: 'too_large.png', size 5.2MB",
        "Upload fails. Client shows error: 'File size exceeds maximum limit of 5MB.'",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of empty (0-byte) JPG file",
        "User is logged in",
        "1. Select a 0-byte file named 'empty_form.jpg'\n2. Click 'Upload'",
        "File: 'empty_form.jpg' (0 bytes)",
        "Upload fails. Error message: 'Cannot upload an empty file.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of corrupted PDF file",
        "User is logged in",
        "1. Select a corrupted PDF file\n2. Click 'Upload'",
        "File: 'corrupt.pdf'",
        "Upload fails. Error message: 'Invalid or corrupted PDF file.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of unsupported format (.txt)",
        "User is logged in",
        "1. Select a text file\n2. Click 'Upload'",
        "File: 'notes.txt'",
        "Upload blocked. Error message: 'File format not supported. Only JPG, PNG, and PDF are allowed.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of unsupported format (.zip)",
        "User is logged in",
        "1. Select a ZIP file\n2. Click 'Upload'",
        "File: 'forms.zip'",
        "Upload blocked. Error: 'File format not supported. Only JPG, PNG, and PDF are allowed.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of unsupported format (.docx)",
        "User is logged in",
        "1. Select a MS Word document file\n2. Click 'Upload'",
        "File: 'form.docx'",
        "Upload blocked. Error: 'File format not supported. Only JPG, PNG, and PDF are allowed.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify file upload with double extension (.pdf.jpg)",
        "User is logged in",
        "1. Select file 'test.pdf.jpg'\n2. Click 'Upload'",
        "File: 'test.pdf.jpg'",
        "Server validates file mime-type. If image payload is clean, uploads; otherwise blocks.",
        "Medium", "Minor", "Security", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify file upload with special characters in name",
        "User is logged in",
        "1. Select file named 'form#%^&@_123.pdf'\n2. Click 'Upload'",
        "File: 'form#%^&@_123.pdf'",
        "Form uploaded successfully. Filename sanitized on the server before storage.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify file upload with extremely long name (250+ characters)",
        "User is logged in",
        "1. Select file with extremely long name\n2. Click 'Upload'",
        "File: 'a'*250 + '.jpg'",
        "Form uploaded successfully. Filename truncated or hashed on server to prevent database overflow.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify real-time upload progress bar indicator accuracy",
        "User is logged in and selects a 4MB file",
        "1. Tap 'Upload'\n2. Monitor progress bar during transmission",
        "File: 4MB JPG file",
        "Progress bar fills dynamically from 0% to 100% in sync with actual upload progress.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify aborting of file upload before completion",
        "User is uploading a large PDF form",
        "1. Click the 'Cancel' or 'X' button during upload progress",
        "File: 4.5MB PDF file",
        "Upload terminates immediately. Incomplete file deleted from backend storage.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload retry logic on network disconnect",
        "User is uploading a file",
        "1. Disconnect internet network connection mid-upload\n2. Observe client response",
        "N/A",
        "Upload fails. Client displays: 'Network disconnected. Tap to retry.' Upload resumes from 0% or cached chunk.",
        "Medium", "Major", "Functional", "No", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of password-protected PDF form",
        "User is logged in",
        "1. Select password-protected PDF file\n2. Click 'Upload'",
        "File: 'protected.pdf' (encrypted)",
        "Upload fails or prompts for password. Error message: 'This PDF is password-protected.'",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify file upload while user is completely offline",
        "User is offline",
        "1. Disconnect network\n2. Click 'Choose File' & select file\n3. Click 'Upload'",
        "N/A",
        "Upload button disabled or error triggers: 'Internet connection required to upload documents.'",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify file records saved in Database",
        "User uploads document successfully",
        "1. Verify database state for `documents` table",
        "N/A",
        "Database contains record with valid user_id, file_path, upload_timestamp, and status 'Pending'.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of duplicate form files in sequence",
        "User has uploaded 'form.jpg'",
        "1. Select the same 'form.jpg' file\n2. Click 'Upload'",
        "File: 'form.jpg'",
        "File uploaded successfully. Server appends unique identifier/timestamp to differentiate them.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload API rate limiting configuration",
        "User has just uploaded 5 forms in 10 seconds",
        "1. Try uploading a 6th form file quickly",
        "File: 'sbi_form.jpg'",
        "Request blocked. HTTP 429 Too Many Requests response. Error: 'Upload limit exceeded. Try again in a minute.'",
        "High", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify bypass security check for executable renamed to PDF",
        "User changes extension of file from 'script.exe' to 'script.pdf'",
        "1. Select 'script.pdf'\n2. Click 'Upload'",
        "File: 'script.pdf' (renamed exe)",
        "Server executes MIME type verification by reading magic bytes, rejecting file as invalid PDF.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of PNG form with transparent alpha layer",
        "User selects transparent PNG template",
        "1. Upload file\n2. Run OCR processing",
        "File: 'transparent_form.png'",
        "Upload and OCR complete successfully. Transparency doesn't interfere with field reading.",
        "Low", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Form Upload", "Upload Form", "Verify upload of highly blurry scanned PDF document",
        "User selects scanned PDF form",
        "1. Upload PDF form file\n2. Run OCR processing",
        "File: 'blurry_scanned.pdf'",
        "Upload succeeds. OCR returns warning about low accuracy due to page quality.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),

    # E. Guidance Testing (25 Test Cases)
    (
        "Guidance", "Field Detection", "Verify guidance generation for standard Bank account form",
        "User is logged in, has processed a standard SBI bank opening form",
        "1. View extracted guidance overlay",
        "SBI Account Opening Form",
        "Fields (e.g. Account type, Name, PAN) detected, labeled, and detailed guidance instructions shown.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance generation for Aadhar registration form",
        "User has uploaded standard Aadhar form",
        "1. View guidance overlays",
        "Aadhar Registration Form",
        "Fields like 'Resident Type', 'Biometrics', 'Parent Name' highlighted with correct filling guidelines.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance generation for PAN card application form",
        "User has processed PAN form",
        "1. View guidance details",
        "PAN Card Form 49A",
        "Fields like 'Assessing Officer Code', 'Identity Proof' highlighted with specific codes guidance.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify 'Full Name' field highlight and tooltips",
        "User is in Guidance screen",
        "1. Tap on the highlighted box for 'Full Name' field",
        "N/A",
        "Tooltip pops up: 'Enter your name exactly as it appears on your ID proof (Aadhar/PAN).'",
        "High", "Major", "UI", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify 'Date of Birth' field validation format guidance",
        "User is in Guidance screen",
        "1. Tap on the 'Date of Birth' highlighted box",
        "N/A",
        "Tooltip: 'Use DD/MM/YYYY format. For example, 15/08/1947.'",
        "High", "Major", "UI", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify 'Signature' box guidance annotation",
        "User is in Guidance screen",
        "1. Scroll to signature section and tap highlight box",
        "N/A",
        "Tooltip: 'Sign only within this boundary. Do not touch the box margins.'",
        "High", "Major", "UI", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify 'Address' fields guidance instructions",
        "User is in Guidance screen",
        "1. Tap highlighted 'Address' box",
        "N/A",
        "Tooltip: 'Enter full postal address. Ensure PIN code is correct for dispatch.'",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify optional fields do not display validation errors",
        "Form contains optional 'Alternate Email' field",
        "1. Inspect highlight color and guidance style for optional fields",
        "N/A",
        "Optional field shown with different color frame (e.g. blue instead of green/red) indicating optional.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify blank mandatory field alert indicator",
        "User has skipped filling a mandatory field during guidance run",
        "1. Try completing the wizard without visiting mandatory field highlight",
        "N/A",
        "Wizard displays a warning or shifts focus back to the missed mandatory field.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify English guidance texts language consistency",
        "User selected English language in application",
        "1. View guidance card text on PAN Form",
        "Language: English",
        "All instructions are in clean English without spelling or translation bugs.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify Telugu guidance texts language consistency",
        "User selected Telugu language in profile",
        "1. View guidance card text on SBI form",
        "Language: Telugu",
        "All guidance card descriptions are successfully translated into correct grammatical Telugu text.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify Hindi guidance texts language consistency",
        "User selected Hindi language in profile",
        "1. View guidance card text on SBI form",
        "Language: Hindi",
        "All guidance card descriptions are translated into proper Hindi text.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify Tamil guidance texts language consistency",
        "User selected Tamil language in profile",
        "1. View guidance card text on SBI form",
        "Language: Tamil",
        "All guidance card descriptions are translated into proper Tamil text.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance overlay mapping scales properly with pinch zoom",
        "User is on the Form Guidance screen",
        "1. Pinch-to-zoom into the form page\n2. Verify coordinates alignment",
        "N/A",
        "Overlay highlighted bounding boxes scale and shift coordinates in sync with the underlying image.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify tap on highlighted field updates details pane content",
        "User is in Guidance screen",
        "1. Tap Field A\n2. Observe details pane\n3. Tap Field B\n4. Observe details pane",
        "Field A (Name), Field B (DOB)",
        "Details pane dynamically updates information to match the currently selected field.",
        "High", "Major", "UI", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance for multi-page forms pagination flow",
        "User uploads a 3-page PDF form",
        "1. Open form guidance wizard\n2. Navigate through steps on page 1\n3. Advance to page 2",
        "Multi-page form",
        "App shifts the background document view to page 2, rendering correct bounding boxes for page 2.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance engine error fallback when parser fails",
        "OCR text is retrieved but templates engine cannot match form layout",
        "1. Open guidance screen for unknown form type",
        "File: 'unknown_form.jpg'",
        "Displays fallback alert: 'Layout not recognized. Please fill fields standardly.' Generates generic fields guidance.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify step-by-step navigation buttons ('Next' / 'Previous')",
        "User is in Guidance screen",
        "1. Tap 'Next' button\n2. Tap 'Previous' button",
        "N/A",
        "Selection moves forward and backward sequentially through the numbered fields.",
        "High", "Major", "UI", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance wizard load speed",
        "User opens form guidance screen",
        "1. Start timer on clicking 'Generate Guidance' until rendering completes",
        "N/A",
        "Guidance wizard rendered within 3.0 seconds under normal internet speed.",
        "Medium", "Major", "Performance", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify 'Did this help?' thumbs feedback interaction",
        "User is reading guidance card tooltip",
        "1. Tap 'Thumbs Up' icon",
        "Action: Thumbs Up",
        "Count updates in backend, button switches state to selected, success confirmation toast displayed.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance info limits for numeric input fields",
        "User taps on 'Aadhar Number' field highlight",
        "1. Review text guidance",
        "Aadhar Field",
        "Tooltip details limit: 'Enter your 12-digit UID number. No spaces or dashes.'",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance text compatibility with standard date format variations",
        "User is viewing date guidance",
        "1. Check if both DD-MM-YYYY and YYYY-MM-DD instructions are provided where appropriate",
        "N/A",
        "Guidance customizes to match formatting of that specific country/template format.",
        "Low", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify overlay alignment doesn't drift on screen orientation change",
        "User rotates mobile device while in Guidance wizard",
        "1. Rotate from portrait to landscape",
        "Screen orientation toggle",
        "Bounding box frames recalculate coordinates and align perfectly on the rotated image.",
        "Medium", "Major", "UI", "No", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify field highlight bounds are legible on low-contrast forms",
        "User uploads low contrast scanner document",
        "1. View guidance overlays",
        "Low contrast form",
        "Highlights maintain high visibility stroke thickness and distinct contrast color (e.g. bright orange).",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Guidance", "Field Detection", "Verify guidance history records list updates",
        "User completes a guidance walkthrough",
        "1. Navigate back to Home screen\n2. Open 'Form History' list",
        "N/A",
        "The completed form appears as the latest entry in the user's form history list.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),

    # F. Voice Guidance Testing (25 Test Cases)
    (
        "Voice Guidance", "Audio Generation", "Verify English TTS audio generation for form fields",
        "User is on the Form Guidance page with English selected",
        "1. Click the 'Voice' icon on a field card",
        "Language: English, Field: 'Full Name'",
        "FastAPI generates and streams correct English voice audio guide files.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Voice Guidance", "Audio Generation", "Verify Telugu TTS audio generation for form fields",
        "User has Telugu selected",
        "1. Click the 'Voice' icon on a field card",
        "Language: Telugu, Field: 'Full Name'",
        "FastAPI/TTS generates and streams correct Telugu voice guide.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Voice Guidance", "Audio Generation", "Verify Hindi TTS audio generation for form fields",
        "User has Hindi selected",
        "1. Click 'Voice' icon",
        "Language: Hindi, Field: 'Address'",
        "FastAPI/TTS streams correct Hindi voice guide.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Voice Guidance", "Audio Generation", "Verify Tamil TTS audio generation for form fields",
        "User has Tamil selected",
        "1. Click 'Voice' icon",
        "Language: Tamil, Field: 'Address'",
        "FastAPI/TTS streams correct Tamil voice guide.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify audio playback play command controls",
        "Voice player is visible in guidance bar",
        "1. Tap 'Play' button",
        "N/A",
        "Audio voice guide begins playing; play button toggles to pause state.",
        "High", "Major", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify audio playback pause control",
        "Audio is currently playing",
        "1. Tap 'Pause' button",
        "N/A",
        "Audio pauses instantly; playback state/position is retained.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify audio playback resume control",
        "Audio is paused",
        "1. Tap 'Play/Resume' button",
        "N/A",
        "Audio resume playback from the exact position where it was paused.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify audio playback stop control",
        "Audio is playing or paused",
        "1. Tap the 'Stop' button",
        "N/A",
        "Audio playing terminates. Progress seeker bar resets to 00:00.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify fast-forward audio track command",
        "Audio is playing",
        "1. Tap 'Forward 10s' button",
        "N/A",
        "Playback position skips forward by 10 seconds or to the end.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify rewind audio track command",
        "Audio is playing",
        "1. Tap 'Rewind 10s' button",
        "N/A",
        "Playback position skips backward by 10 seconds or to the start.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify speed controls adjustment options",
        "Audio is playing",
        "1. Select speed multiplier from options (0.75x, 1.25x, 1.5x)",
        "Speed: 1.5x",
        "Audio plays faster without pitch distortion.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify device volume control sync",
        "Audio is playing",
        "1. Adjust device physical volume buttons up and down",
        "Physical keys press",
        "Volume of voice guide audio scales in alignment with system media volume.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify mute and unmute buttons functionality",
        "Audio is playing",
        "1. Click the 'Mute' speaker icon\n2. Click it again to unmute",
        "N/A",
        "Audio turns silent instantly; unmuting restores volume to original state.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Language Switching", "Verify instant language switch during audio playback",
        "Audio is playing in English",
        "1. Open language dropdown\n2. Select Telugu",
        "Selected: Telugu",
        "English playback stops immediately, and Telugu voice track starts playing for the same field.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify audio playback state background safety",
        "Audio is playing in-app",
        "1. Minimize application to background\n2. Open device notification drawer",
        "N/A",
        "Audio continues playing or pauses based on configuration; notification controller is shown.",
        "Medium", "Minor", "Functional", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify audio playback interruption handling by incoming call",
        "Audio is playing",
        "1. Simulate incoming telephone call",
        "N/A",
        "App voice audio pauses automatically. Resumes or remains paused after call is completed.",
        "High", "Major", "Functional", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify headset disconnection behavior during playback",
        "Audio is playing through connected headphones (Bluetooth/Jack)",
        "1. Disconnect headphone device",
        "N/A",
        "Playback pauses automatically to prevent audio routing to speaker.",
        "Medium", "Minor", "Functional", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Generation", "Verify Kotlin audio player format compatibility",
        "API returns voice file bytes",
        "1. Inspect streamed file content type",
        "Content-Type: audio/mpeg",
        "Response is clean MP3 stream compatible with Jetpack Compose MediaPlayer/ExoPlayer.",
        "High", "Major", "Integration", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify offline voice guide warning display",
        "Device is offline",
        "1. Attempt to play voice guide of a field not locally cached",
        "N/A",
        "Error tooltip: 'Audio guides require network connection. Please connect online.'",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Generation", "Verify voice gender switching profile options",
        "User is in Settings profile screen",
        "1. Toggle voice choice preference from Female to Male\n2. Play field voice guide",
        "Voice setting: Male",
        "The generated audio uses a male voice model.",
        "Low", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify loading progress spinner for slow voice buffers",
        "User is on weak network (2G)",
        "1. Click 'Play Voice'\n2. Observe buffer behavior",
        "Weak connection speed",
        "Buffer spinner displays on player bar until first batch of audio chunks is cached.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Generation", "Verify backend TTS engine fallback when API limit is hit",
        "Third party TTS API limits exhausted",
        "1. Trigger high count of voice generation requests",
        "N/A",
        "Backend falls back to local Android TTS configuration or displays generic offline voice warning.",
        "High", "Major", "Integration", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify visual highlight sync with voice reader",
        "Voice guide is playing step-by-step automatically",
        "1. Observe active field highlights as audio instructions read through",
        "Auto-play mode active",
        "The red border overlay focus moves automatically to the next field as the voice reader finishes.",
        "High", "Major", "UI", "No", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify volume level limits on high density forms",
        "User is reading large bank document guide",
        "1. Rapidly switch field selection back and forth while playing audio",
        "N/A",
        "Current audio thread terminates cleanly and the new field's audio starts without overlapping.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify cleanup of temporary downloaded audio caches on logout",
        "User has played 10 audio guides (cached in device app directory)",
        "1. Tap 'Logout'\n2. Inspect app cache directory storage",
        "N/A",
        "All temporary downloaded .mp3 voice guide cache files are deleted from the device.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),

    # G. Database Testing (25 Test Cases)
    (
        "Database", "User Records", "Verify user details row insertion on successful signup",
        "User registered",
        "1. Connect to Railway/Aiven MySQL database\n2. Query `users` table for email",
        "Email: 'ramesh.g@example.com'",
        "Record exists with correct username, status active, and password containing hash string.",
        "High", "Critical", "Database", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Database", "User Records", "Verify passwords stored as bcrypt hashes in database",
        "User registered",
        "1. Read `password_hash` column for registered user",
        "N/A",
        "Password is not stored as plain text. Format matches '$2b$12$...' indicating bcrypt.",
        "High", "Critical", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify DB constraints block duplicate emails",
        "Database is running",
        "1. Attempt direct insert SQL query with existing email",
        "INSERT query with duplicate email",
        "MySQL returns error code 1062 (Duplicate entry for key 'email'). Transaction rolls back.",
        "High", "Critical", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify profile field changes reflect in database immediately",
        "User changed profile name",
        "1. Query `users` table for the modified row",
        "N/A",
        "Row columns match updated entries instantly.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify profile image URL path is correctly written to DB",
        "User uploaded new profile image avatar",
        "1. Query user record in DB",
        "N/A",
        "Column `profile_image_url` updated with path to CDN/Render host file storage.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify default language setting preference matches English in DB",
        "New user registration",
        "1. Read `users.language_pref` column values",
        "N/A",
        "Defaults to 'en' (English) upon creation.",
        "Medium", "Minor", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "Documents Records", "Verify uploaded document metadata insertion in documents table",
        "User uploaded form document successfully",
        "1. Query database `documents` table",
        "N/A",
        "Row successfully created with columns: id, user_id, filename, file_path, upload_time, status.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "Documents Records", "Verify OCR text cache storage after processing",
        "EasyOCR completed text extraction on document ID 45",
        "1. Check table `ocr_cache` or `document_guidance` in DB",
        "Document ID: 45",
        "Extracted text data payload stored correctly in JSON/TEXT column fields.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "Documents Records", "Verify OCR bounding box JSON coordinates structure in database",
        "OCR data saved",
        "1. Inspect coordinate schema layout in DB column",
        "N/A",
        "JSON matches schema format structure: {'box_id': int, 'coords': [x, y, w, h]}.",
        "Medium", "Minor", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify auto-incrementing key generation of primary keys",
        "Inserting multiple records in sequence",
        "1. Check primary keys sequence indexes",
        "N/A",
        "IDs increment linearly (1, 2, 3...) without duplicates or jumps.",
        "Medium", "Minor", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "Feedback Records", "Verify feedback submission details write",
        "User submitted feedback",
        "1. Query `feedback` table for latest row",
        "N/A",
        "Fields logged: user_id, rating (1-5), message, submission_time.",
        "High", "Major", "Database", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Database", "User Records", "Verify datetime timestamps stored in UTC format",
        "Action logged",
        "1. Retrieve timestamp values from `created_at` in DB",
        "N/A",
        "Timestamp matches ISO UTC format (e.g. 2026-06-23T04:00:00Z).",
        "Medium", "Minor", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify cascade deletion of documents when user is deleted",
        "Admin deletes user ID 10",
        "1. Delete user record with ID 10\n2. Query `documents` table for user_id = 10",
        "User ID: 10",
        "Foreign key constraints execute CASCADE DELETE. All document rows for user 10 are deleted.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "Feedback Records", "Verify cascade deletion of feedback rows when user is deleted",
        "Admin deletes user ID 10",
        "1. Delete user record with ID 10\n2. Query `feedback` table for user_id = 10",
        "User ID: 10",
        "All feedback entries submitted by user 10 are deleted automatically.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify database connection pool handles multiple parallel connections",
        "FastAPI is running",
        "1. Open 50 database sessions simultaneously\n2. Measure connection performance",
        "N/A",
        "FastAPI SQLAlchemy connection pool manages connections smoothly without timeouts.",
        "High", "Major", "Database", "No", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify search index optimizations on indexed search columns",
        "Database contains 50,000 users",
        "1. Run EXPLAIN query on search by email select statement",
        "Query: EXPLAIN SELECT * FROM users WHERE email = 'ramesh.g@example.com'",
        "Query utilizes index key. Key scans are minimized.",
        "Medium", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify rollback of transactions during mid-operation API faults",
        "API fails during complex double-table update",
        "1. Trigger partial user profile update that fails on second query",
        "N/A",
        "First query changes are rolled back. Database remains in consistent state.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify data encryption parameters for user tables sensitive columns",
        "Inspect database configuration",
        "1. Check if tables are encrypted at rest (AES-256)",
        "N/A",
        "Railway/Aiven MySQL tables are configured with at-rest encryption.",
        "Medium", "Minor", "Database", "No", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify database backup recovery script checks",
        "Automated backup triggers daily",
        "1. Restore database from backup SQL dump\n2. Verify checksum integrity",
        "Backup SQL file",
        "Restored database is identical to primary DB. Schema and indexes verify.",
        "High", "Critical", "Database", "No", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify foreign key constraint error when deleting parent records with restrict",
        "User has processed documents",
        "1. Attempt deleting user without cascading constraints configured (e.g. key tables with RESTRICT rule)",
        "N/A",
        "MySQL blocks delete query, throws foreign key violation error.",
        "Medium", "Minor", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify Alembic schema migration updates execution integrity",
        "New column added in local development",
        "1. Run alembic upgrade head script\n2. Check column existence in target DB",
        "N/A",
        "Schema updates successfully without corrupting existing records.",
        "High", "Major", "Database", "No", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify database handles emoji characters input parameters",
        "User enters name with emojis",
        "1. Update user name in DB\n2. Query name values",
        "Name: 'Ramesh 🇮🇳'",
        "DB columns are configured with utf8mb4 encoding, name saves and reads correctly.",
        "Medium", "Minor", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify connection timeouts constraints parameters in FastAPI connection pool",
        "Database server goes down",
        "1. Disconnect MySQL server instance\n2. Call user dashboard API",
        "N/A",
        "FastAPI returns connection error within 5 seconds instead of hung thread.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "Documents Records", "Verify form guidance history table query schema mapping",
        "User requests list of guidance runs",
        "1. Query `guidance_runs` table",
        "N/A",
        "Returns correct mapping link between user_id, document_id, template_id, status.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),
    (
        "Database", "User Records", "Verify data replica replication lag checks",
        "Write query executed on primary DB",
        "1. Immediately read query from replica database",
        "N/A",
        "Replication lag is below 500ms; user sees updated state immediately.",
        "Low", "Minor", "Database", "No", ["Regression"]
    ),

    # H. API Testing (25 Test Cases)
    (
        "API", "Authentication API", "Verify signup API response on success",
        "FastAPI is running and database connection is healthy",
        "1. Send POST request to `/api/auth/signup` with unique payload JSON",
        "JSON payload: username, email, password",
        "Returns HTTP 201 Created and JSON response: {'status': 'success', 'message': 'User registered'}.",
        "High", "Critical", "API", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "API", "Authentication API", "Verify login API response on success",
        "User is registered",
        "1. Send POST to `/api/auth/login` with valid email & password",
        "JSON payload: email, password",
        "Returns HTTP 200 OK and access token dictionary containing JWT.",
        "High", "Critical", "API", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "API", "User Profile API", "Verify retrieve user profile API response schema",
        "User is authenticated",
        "1. Send GET to `/api/user/profile` with valid authorization Bearer token",
        "Header: Authorization: Bearer <token>",
        "Returns HTTP 200 OK and user information JSON structure matching profile schema.",
        "High", "Critical", "API", "Yes", ["Smoke", "Regression"]
    ),
    (
        "API", "Documents API", "Verify form upload API response schema",
        "User is authenticated",
        "1. Send POST to `/api/documents/upload` containing multipart file payload",
        "Header: Bearer token, Payload: form.jpg file",
        "Returns HTTP 201 Created and response: {'document_id': int, 'status': 'uploaded'}.",
        "High", "Critical", "API", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "API", "Documents API", "Verify form guidance generation API response schema",
        "Document is uploaded and processed",
        "1. Send GET to `/api/documents/{id}/guidance` with document ID",
        "Document ID: 15",
        "Returns HTTP 200 OK with list of guidance fields containing bounding boxes and steps descriptions.",
        "High", "Critical", "API", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "API", "Documents API", "Verify voice guidance audio streaming API response",
        "Audio is generated on server",
        "1. Send GET to `/api/documents/{id}/audio` with field index and language params",
        "ID: 15, field: 3, lang: 'te'",
        "Returns HTTP 200 OK and streams audio/mpeg binary file data directly.",
        "High", "Major", "API", "Yes", ["Smoke", "Regression"]
    ),
    (
        "API", "Feedback API", "Verify submit feedback API response schema",
        "User is authenticated",
        "1. Send POST to `/api/feedback` with rating and text payload",
        "Payload: rating: 5, message: 'Excellent assistance!'",
        "Returns HTTP 201 Created and feedback confirmed status JSON.",
        "Medium", "Major", "API", "Yes", ["Sanity", "Regression"]
    ),
    (
        "API", "Admin API", "Verify admin retrieve users list API response schema",
        "User is logged in as Admin role",
        "1. Send GET to `/api/admin/users` with pagination limit parameters",
        "Parameters: page: 1, limit: 10",
        "Returns HTTP 200 OK and paginated lists containing user array, total count, pages count.",
        "High", "Major", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Authentication API", "Verify signup API response for duplicate email error",
        "Email exists in database",
        "1. Send POST to `/api/auth/signup` with duplicate email address",
        "Payload containing existing email",
        "Returns HTTP 400 Bad Request with error: 'Email address already registered.'",
        "High", "Major", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Authentication API", "Verify login API response for invalid password",
        "User exists in system",
        "1. Send POST to `/api/auth/login` with incorrect password details",
        "Payload with incorrect password",
        "Returns HTTP 401 Unauthorized with detail message: 'Invalid email or password.'",
        "High", "Major", "API", "Yes", ["Sanity", "Regression"]
    ),
    (
        "API", "User Profile API", "Verify API blocks profile requests when authorization header is missing",
        "User session token is cleared",
        "1. Send GET request to `/api/user/profile` without Authorization header",
        "N/A",
        "Returns HTTP 401 Unauthorized with detail: 'Not authenticated.'",
        "High", "Critical", "API", "Yes", ["Smoke", "Regression"]
    ),
    (
        "API", "User Profile API", "Verify API blocks request when token signature is invalid",
        "User sends tampered token",
        "1. Send GET request with modified Bearer token signature payload",
        "Header: Authorization: Bearer <tampered_token>",
        "Returns HTTP 403 Forbidden or 401 Unauthorized with detail: 'Invalid token.'",
        "High", "Critical", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Documents API", "Verify documents API response for non-existent document ID",
        "User is authenticated",
        "1. Send GET request to `/api/documents/999999/guidance` path",
        "Document ID: 999999",
        "Returns HTTP 404 Not Found with error detail: 'Document not found.'",
        "High", "Major", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Documents API", "Verify document upload API error response when file missing in payload",
        "User is authenticated",
        "1. Send POST request to `/api/documents/upload` with empty multipart body",
        "Payload: empty request body",
        "Returns HTTP 422 Unprocessable Entity specifying details of missing parameter 'file'.",
        "High", "Major", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Integration", "Verify API behavior when MySQL database server is offline",
        "Database server stopped",
        "1. Send GET request to `/api/user/profile` with valid authorization token",
        "N/A",
        "Returns HTTP 500 Internal Server Error, masking details of database crash.",
        "High", "Critical", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Integration", "Verify CORS headers are returned in API responses",
        "FastAPI running",
        "1. Send OPTIONS request to `/api/auth/login` from allowed origins list",
        "Headers: Origin: http://localhost:5173",
        "Response headers contain: Access-Control-Allow-Origin: http://localhost:5173 and allowed methods.",
        "High", "Major", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Integration", "Verify response compression compression headers",
        "User requests large data array",
        "1. Send GET request to `/api/admin/users` with compression headers",
        "Header: Accept-Encoding: gzip",
        "Response header contains Content-Encoding: gzip and response data is compressed.",
        "Medium", "Minor", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Security", "Verify API rate limit limits are enforced",
        "Rate limiter configured for 60 requests per minute",
        "1. Send 65 rapid GET requests to `/api/user/profile` in 10 seconds",
        "N/A",
        "Returns HTTP 429 Too Many Requests after 60 requests. Retry header specifies sleep time.",
        "High", "Major", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Integration", "Verify api behavior on connection loss mid-request",
        "Upload in progress",
        "1. Force disconnect client connection during upload stream transmission",
        "N/A",
        "Server handles connection drop gracefully without memory leaks or leaving partial files.",
        "High", "Major", "API", "No", ["Regression"]
    ),
    (
        "API", "User Profile API", "Verify PUT profile API updates user records and returns success schema",
        "User is logged in",
        "1. Send PUT request to `/api/user/profile` with updated name and preferences",
        "JSON payload: name: 'Updated Name', lang: 'te'",
        "Returns HTTP 200 OK and response updates data values in return payload.",
        "High", "Major", "API", "Yes", ["Sanity", "Regression"]
    ),
    (
        "API", "Documents API", "Verify DELETE documents API drops files and DB rows",
        "User is logged in and has uploaded files",
        "1. Send DELETE request to `/api/documents/15`",
        "Document ID: 15",
        "Returns HTTP 200 OK. Subsequent GET to same URL returns 404.",
        "High", "Major", "API", "Yes", ["Sanity", "Regression"]
    ),
    (
        "API", "Performance", "Verify API response latency threshold",
        "Server is running normal operations load",
        "1. Send standard GET request to `/api/auth/health` endpoint\n2. Log execution duration",
        "N/A",
        "API response duration is below 200ms.",
        "Medium", "Minor", "API", "Yes", ["Regression"]
    ),
    (
        "API", "Security", "Verify API response on oversized payload JSON injection attempt",
        "API is configured with max payload body checks",
        "1. Send POST request with 50MB JSON payload",
        "Payload: 50MB random text",
        "Returns HTTP 413 Payload Too Large and terminates connection immediately.",
        "High", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "API", "Security", "Verify API redirects HTTP requests to HTTPS endpoint",
        "Server running production configuration on Render",
        "1. Send HTTP request to http://formsahayak-api.render.com/api/user/profile",
        "N/A",
        "Response status code 301 Moved Permanently redirects request to HTTPS location.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "API", "Authentication API", "Verify token refresh endpoint validation rules",
        "JWT access token is expired, refresh token remains valid",
        "1. Send POST to `/api/auth/refresh` containing refresh token payload",
        "JSON containing valid refresh token",
        "Returns HTTP 200 OK and generates new JWT access token strings.",
        "High", "Major", "API", "Yes", ["Regression"]
    ),

    # I. UI Testing (25 Test Cases)
    (
        "UI", "Buttons", "Verify registration button changes states visually",
        "User is on the Registration screen",
        "1. Tap and hold 'Sign Up' button\n2. Release button",
        "N/A",
        "Button background changes color on active tap, showing responsive tactile feedback.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Navigation", "Verify bottom navigation bar switches dashboard fragments",
        "User is logged into application dashboard",
        "1. Tap 'Upload Form' icon\n2. Tap 'History' icon\n3. Tap 'Settings' icon",
        "N/A",
        "Dashboard switches views smoothly to the corresponding tab fragments.",
        "High", "Major", "UI", "No", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "UI", "Responsiveness", "Verify UI layout rendering on 5.5-inch mobile screen",
        "App is running on standard smartphone device",
        "1. View login page and forms dashboard",
        "Screen size: 5.5 inches",
        "Layout adjusts correctly. No text overflows, overlaps, or cropped buttons.",
        "High", "Major", "UI", "No", ["Sanity", "Regression"]
    ),
    (
        "UI", "Responsiveness", "Verify UI layout rendering on 10-inch tablet screen",
        "App is running on standard tablet device",
        "1. View navigation bar and dashboard panel",
        "Screen size: 10 inches",
        "UI elements scale proportionally, utilizing extra grid space correctly.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Responsiveness", "Verify UI layout adjustments in portrait mode",
        "App is active on standard smartphone",
        "1. Set screen orientation to portrait",
        "Orientation: Portrait",
        "View layouts align vertically, bottom navbar displays at screen base.",
        "High", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Responsiveness", "Verify UI layout adjustments in landscape mode",
        "App is active on standard smartphone",
        "1. Rotate screen orientation to landscape",
        "Orientation: Landscape",
        "Layout displays with side menu scrollable or elements reflowed correctly.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Dark Mode", "Verify color contrast and text readability in Dark Mode",
        "App is running on system with dark mode activated",
        "1. Switch app settings to Dark Mode\n2. Review text on all screens",
        "Setting: Dark Mode",
        "Background turns dark gray/black. Text colors shift to white/gray, maintaining contrast ratio >= 4.5:1.",
        "High", "Major", "UI", "No", ["Sanity", "Regression"]
    ),
    (
        "UI", "Dark Mode", "Verify color contrast and text readability in Light Mode",
        "App is running with light mode activated",
        "1. Switch app settings to Light Mode\n2. Review screen elements",
        "Setting: Light Mode",
        "Background turns light gray/white. Text elements display dark gray/black.",
        "High", "Major", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify text input fields focus border colors states",
        "User is on the login screen",
        "1. Tap inside 'Email' text input box",
        "N/A",
        "Input boundary highlights in distinct color (e.g. primary teal), indicating active focus state.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify red error alerts display below invalid fields",
        "User submits invalid registration details",
        "1. Tap 'Sign Up'",
        "N/A",
        "Validation error texts render in small red font directly below each invalid field.",
        "High", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Accessibility", "Verify screen reader accessibility compatibility",
        "Device TalkBack / voice assistant is activated",
        "1. Navigate through dashboard using swipe gestures",
        "TalkBack mode active",
        "Focus squares select each button, reading aloud content description metadata.",
        "Medium", "Minor", "Accessibility", "No", ["Regression"]
    ),
    (
        "UI", "Accessibility", "Verify dynamic font size scaling rendering safety",
        "Device settings font size set to Extra Large",
        "1. Open dashboard screen\n2. Inspect labels",
        "Font scaling: 150%",
        "Labels grow in size, wraps text lines correctly without clipping inside container boxes.",
        "Medium", "Minor", "Accessibility", "No", ["Regression"]
    ),
    (
        "UI", "Responsiveness", "Verify UI layouts handle long text localization strings",
        "App language set to German or Tamil which has long character sequences",
        "1. Navigate to Settings page options",
        "Language: Tamil",
        "Button text adapts or wraps to double lines cleanly without overlapping margins.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Navigation", "Verify side navigation drawer drawer slides cleanly",
        "App dashboard layout uses side drawer menu navigation",
        "1. Swipe right from screen edge or tap hamburger menu icon",
        "N/A",
        "Navigation drawer slides open smoothly from left; background dims under translucent overlay.",
        "High", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Navigation", "Verify image viewport pinch-to-zoom controls navigation controls",
        "User is viewing form guidance page",
        "1. Zoom in on image\n2. Pan around the form dimensions using drag gestures",
        "N/A",
        "Viewport pans smoothly, keeping bounding box focus outlines aligned over targets.",
        "High", "Major", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Navigation", "Verify device hardware back button navigation stack order",
        "User navigated: Login -> Language -> Dashboard -> Settings",
        "1. Press hardware back button\n2. Press back button again",
        "N/A",
        "App screen returns back to Dashboard, then exits to home or locks screen stack.",
        "High", "Major", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify auto keyboard pop-up behavior inside search fields",
        "User clicks search bar icon in admin panel",
        "1. Tap the search input area",
        "N/A",
        "Software virtual keyboard slides open from screen base immediately.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify keyboard action buttons toggle states in registration forms",
        "User is filling username field in signup",
        "1. Review bottom right action button on software keyboard",
        "N/A",
        "Keyboard displays 'Next' button, shifting focus directly to Email field upon tap.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify skeleton screen shimmer displays during api wait times",
        "User loaded form history with slow network speed",
        "1. Open History list view",
        "N/A",
        "Grey pulsing box lines display placeholder content while API query loads.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify button disable states during API loading times",
        "User clicked 'Upload File' submit action",
        "1. Monitor button visibility during upload transmission",
        "N/A",
        "Button text changes to 'Uploading...' or loading circle spins; clicks are blocked.",
        "High", "Major", "UI", "No", ["Smoke", "Regression"]
    ),
    (
        "UI", "Buttons", "Verify visual alignment consistency of alert model windows",
        "Alert popup triggers in-app",
        "1. Review layout borders",
        "N/A",
        "Alert window centers perfectly with standardized padding margins.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify tooltips behavior on hover or long-press",
        "User is looking at help icons",
        "1. Long press (?) question mark indicator on form field",
        "N/A",
        "Small helper tooltip container box pops open, closing on touch release.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Responsiveness", "Verify app layout rendering safety for screens with notches/punch-holes",
        "Running app on phone with camera punch hole overlay",
        "1. Open settings drawer\n2. Inspect top margin spacing",
        "Screen size: standard notch phone",
        "App title and layout contents are pushed down below status bar height margins.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Navigation", "Verify transition animations smoothness between screens",
        "User shifts views",
        "1. Navigate between Dashboard page and About Developer screen",
        "N/A",
        "Views fade or slide with smooth transitions without UI stuttering.",
        "Medium", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "UI", "Buttons", "Verify high resolution rendering of logos across screen densities",
        "App running on high density screen device",
        "1. Inspect branding logo on splash login screen",
        "Screen size: QHD display",
        "Logo graphics look sharp. No visible pixelation borders.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),

    # J. Security Testing (25 Test Cases)
    (
        "Security", "SQL Injection", "Verify SQL injection protection in login inputs",
        "User is on the login page",
        "1. Enter `' OR 1=1; --` in Email field\n2. Enter random password\n3. Click 'Login'",
        "Email: 'admin@test.com\' OR 1=1; --'",
        "Authentication fails. Input query handles parameters safely without executing injected SQL.",
        "High", "Critical", "Security", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Security", "SQL Injection", "Verify SQL injection protection in Admin search user query",
        "Admin is logged in and on search user panel",
        "1. Enter query: `' UNION SELECT username, password_hash FROM users; --`\n2. Click search",
        "Query payload",
        "No database columns exposed. Search returns zero results or handles input literally.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "SQL Injection", "Verify SQL injection protection in HTTP request headers",
        "FastAPI receiving requests",
        "1. Send API request containing SQL payload in authorization Bearer token value string",
        "Token: Bearer ' UNION SELECT...",
        "FastAPI middleware validates token integrity, rejecting invalid tokens with 401 Unauthorized.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "XSS Protection", "Verify reflected XSS injection prevention in signup inputs",
        "User registration screen",
        "1. Enter username: `<script>alert('xss')</script>`\n2. Submit registration details",
        "Username payload",
        "Registration fails or database encodes script characters (`&lt;script&gt;`) safely.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "XSS Protection", "Verify stored XSS injection prevention in Feedback form",
        "Feedback submission screen",
        "1. Enter comments: `<img src=x onerror=alert(document.cookie)>`\n2. Submit feedback",
        "Feedback text field payload",
        "Comments are saved. When Admin views this feedback entry on dashboard, no javascript runs.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Unauthorized Access", "Verify unauthorized access blocks to admin endpoints",
        "User is logged in as standard User role",
        "1. Attempt to call admin API endpoint GET `/api/admin/users` in Postman",
        "N/A",
        "Server rejects request with status code 403 Forbidden or 401 Unauthorized.",
        "High", "Critical", "Security", "Yes", ["Smoke", "Regression", "Critical Path"]
    ),
    (
        "Security", "Unauthorized Access", "Verify user cannot access files/metadata of other users",
        "User A and User B exist. User A has document ID 12.",
        "1. User B logs in\n2. User B calls GET `/api/documents/12/guidance`",
        "Document ID: 12",
        "Server returns 403 Forbidden or 404 Not Found, blocking access to cross-user documents.",
        "High", "Critical", "Security", "Yes", ["Smoke", "Regression", "Critical Path"]
    ),
    (
        "Security", "JWT Validation", "Verify JWT signature validation is enforced",
        "User is logged in",
        "1. Extract valid JWT token\n2. Modify data payload (e.g. change role from 'user' to 'admin')\n3. Resubmit API call",
        "Header containing altered JWT",
        "API fails authentication, server identifies signature mismatch and rejects request.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "JWT Validation", "Verify JWT validation algorithm 'none' exploit prevention",
        "User intercepts request",
        "1. Change JWT header key alg to 'none'\n2. Remove signature bytes\n3. Send API call",
        "JWT Header containing alg: 'none'",
        "Server rejects request. FastAPI validation specifies algorithm check and blocks 'none'.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Session Hijacking", "Verify session token invalidation on Logout",
        "User has active session",
        "1. Tap 'Logout' button\n2. Copy the used JWT access token\n3. Attempt to fetch `/api/user/profile` using token",
        "Header containing expired/logged out JWT",
        "Request fails with 401 Unauthorized. Server logs token in blacklist.",
        "High", "Critical", "Security", "Yes", ["Smoke", "Regression"]
    ),
    (
        "Security", "Session Hijacking", "Verify cookies flags configuration",
        "Inspect cookies returned in browser console",
        "1. Inspect HTTP session cookies for web application interface",
        "N/A",
        "Cookies contain security flags: 'HttpOnly', 'Secure', and 'SameSite=Strict'.",
        "Medium", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Cryptography", "Verify password hashing salt iterations count configuration",
        "Review backend FastAPI configuration code",
        "1. Verify salt rounds constant for bcrypt algorithm",
        "N/A",
        "Hash rounds count is configured to minimum of 12 for strong work factor.",
        "Medium", "Minor", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Information Disclosure", "Verify sensitive properties are masked in log outputs",
        "Verify server log files content",
        "1. Review fastapi output logs after user login and update password operations",
        "N/A",
        "Plain text password fields are filtered out and show mask values `*****`.",
        "Medium", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Session Hijacking", "Verify session brute-force limits lock execution",
        "User registers login attempts",
        "1. Send 10 login queries with incorrect password in 1 minute\n2. Send 11th login query",
        "N/A",
        "App blocks requests with HTTP 429 or Locks Account. Alerts user to retry after threshold duration.",
        "High", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Session Hijacking", "Verify executable payload blocking in file uploads",
        "User is on document upload tab",
        "1. Rename file 'virus.exe' to 'virus.pdf'\n2. Click 'Upload'",
        "File: 'virus.pdf'",
        "Server file scanner checks content bytes and rejects upload, preventing arbitrary execution.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Security Policy", "Verify CORS allowed origins is restricted",
        "Review FastAPI configuration",
        "1. Check CORS middleware origin settings",
        "N/A",
        "CORS does not use '*' wildcard in production environment. Specific frontend URLs are configured.",
        "Medium", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Session Hijacking", "Verify Clickjacking protection headers",
        "Inspect dashboard HTTP responses",
        "1. Read returned headers of Admin dashboard requests",
        "N/A",
        "Headers contain X-Frame-Options: DENY or SAMEORIGIN.",
        "Medium", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Security Policy", "Verify SSL/HTTPS redirects are active",
        "API hosted on Render",
        "1. Navigate to endpoint url starting with HTTP://",
        "N/A",
        "API automatically redirects browser request URL port connection to HTTPS secure port.",
        "High", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Unauthorized Access", "Verify BOLA (Broken Object Level Authorization) prevention in profile updates",
        "User A tries to update User B's profile details",
        "1. Send PUT request to `/api/user/profile` containing body with user_id = User B's ID",
        "User A Bearer token, Payload: user_id: 20",
        "Server updates only User A's details based on token metadata, ignoring user_id in body parameters.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Unauthorized Access", "Verify BFLA (Broken Function Level Authorization) prevention",
        "User attempts calling moderator level action",
        "1. Send POST request to `/api/admin/users/disable` with user token credentials",
        "N/A",
        "Server validation blocks request, returning status 403 Forbidden.",
        "High", "Critical", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Session Validation", "Verify password reset validation keys timeout constraints",
        "User requests reset password link",
        "1. Open link after token validation window expires (e.g. 1 hour)",
        "N/A",
        "App displays message: 'Reset token expired. Please request a new password link.'",
        "High", "Major", "Security", "Yes", ["Regression"]
    ),
    (
        "Security", "Cryptography", "Verify database connection SSL encryption rules",
        "Database is hosted on Railway/Aiven cloud MySQL",
        "1. Verify database config strings on FastAPI host",
        "N/A",
        "Database connection uses SSL configuration (ssl_mode=require) preventing sniffed parameters.",
        "High", "Critical", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Security Policy", "Verify proguard obfuscation obfuscation inside Kotlin APK builds",
        "Kotlin App build is packaged",
        "1. Run decompiler tool on built apk",
        "Built APK file",
        "Variables, packages, and classes names are obfuscated (a, b, c). Sensitive API strings are hidden.",
        "Medium", "Minor", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Cryptography", "Verify JWT tokens storage safety inside Android local memory storage",
        "App is running on android device",
        "1. Search local SharedPreferences folder files",
        "N/A",
        "JWT tokens are stored inside EncryptedSharedPreferences container rather than plain text.",
        "High", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Security", "Security Policy", "Verify endpoint resilience against Denial of Service rate floods",
        "API endpoints running",
        "1. Send 1000 requests in 10 seconds via API stress tools",
        "N/A",
        "Rate limiter block and firewall configurations block IP before overloading Render server.",
        "High", "Major", "Security", "No", ["Regression"]
    ),

    # K. Admin Dashboard Testing (25 Test Cases)
    (
        "Admin Dashboard", "Admin Login", "Verify Admin Dashboard login with correct admin credentials",
        "User is registered as Admin role in system DB",
        "1. Enter admin email\n2. Enter correct password\n3. Tap 'Login'",
        "Email: 'admin@formsahayak.com', Password: 'AdminPassword@123'",
        "Redirects user dashboard view immediately to Admin Panel layout controls.",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Admin Dashboard", "Admin Login", "Verify Admin Dashboard login blocked for standard users",
        "User is registered with normal role in database",
        "1. Enter email\n2. Enter password\n3. Click 'Login' as Admin",
        "Email: 'ramesh.g@example.com', Password: 'Password@123'",
        "Login blocked. Error: 'Access denied. Administrator privileges required.'",
        "High", "Critical", "Functional", "Yes", ["Smoke", "Sanity", "Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify complete list of registered users is displayed",
        "Admin is logged in and on the Admin Dashboard panel",
        "1. Click on the 'Users List' side menu option",
        "N/A",
        "Table displays users. Columns show ID, Username, Email, Phone, Language, Date Created, and Status.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify sorting of users list by date created",
        "Admin is viewing the Users List table",
        "1. Click on 'Date Created' column header to sort ascending\n2. Click header again to sort descending",
        "N/A",
        "User rows re-order correctly. Date columns verify order sequence.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify sorting of users list alphabetically",
        "Admin is viewing the Users List table",
        "1. Click on 'Username' column header\n2. Click header again",
        "N/A",
        "User rows sort alphabetically (A-Z) and reverse alphabetically (Z-A).",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Search", "Verify searching users by exact username",
        "Admin is viewing the Users List table. User 'ramesh_g' exists.",
        "1. Type username in search bar\n2. Press Enter or click search icon",
        "Search Query: 'ramesh_g'",
        "Table displays exactly one matching user row details for 'ramesh_g'.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Admin Dashboard", "User Search", "Verify searching users by partial email address",
        "Admin is viewing the Users list. Multiple emails contain '@example.com'.",
        "1. Type partial email domain in search bar\n2. Click search icon",
        "Search Query: '@example.com'",
        "Table displays all user rows containing '@example.com' in email field.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Search", "Verify search results when query matches no records",
        "Admin is viewing the Users List table",
        "1. Type random string in search input\n2. Click search icon",
        "Search Query: 'xyz_non_existent'",
        "Table displays empty rows placeholder. Text: 'No matching users found.'",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify pagination controls on user table list",
        "Database contains 150 registered users",
        "1. View bottom of User table\n2. Click 'Next Page' icon\n3. Click 'Items per page' dropdown & select 50",
        "N/A",
        "Table updates to show correct page offsets. Rows count expands to display 50 users.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify disabling a user account",
        "Admin is viewing user list table details",
        "1. Click the 'Disable/Block' button on User Ramesh row",
        "User ID: 15",
        "User status changes to 'Disabled'. Ramesh session is terminated instantly; logins are blocked.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify deleting a user account",
        "Admin is on the User detail page",
        "1. Click 'Delete User' button\n2. Confirm warning alert modal prompt",
        "User ID: 15",
        "User row is permanently dropped from DB. Cascading drops clear all user documents and history.",
        "High", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "Analytics", "Verify daily user registration counts on Analytics charts",
        "Admin is on the Analytics dashboard tab",
        "1. Review the 'Daily Registrations' line chart visual metrics",
        "N/A",
        "Chart values match actual database counts for each respective calendar date.",
        "Medium", "Major", "Functional", "No", ["Sanity", "Regression"]
    ),
    (
        "Admin Dashboard", "Analytics", "Verify filtering analytics by date range option",
        "Admin is on the Analytics dashboard",
        "1. Click Date filter dropdown\n2. Select 'Last 7 Days'\n3. Select 'Last 30 Days'",
        "N/A",
        "Chart data coordinates recalculate instantly to display metrics inside select window.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "Analytics", "Verify processed forms counter increment updates",
        "Admin is viewing the Analytics counters panels",
        "1. Open user app on Device A and run OCR text guide on one document\n2. Monitor admin page count counter on Device B",
        "N/A",
        "Analytics 'Total Documents Processed' counter increment updates in real-time or page reload.",
        "Medium", "Major", "Functional", "No", ["Regression"]
    ),
    (
        "Admin Dashboard", "Analytics", "Verify server performance stats analytics graphs display",
        "Admin is on the System Health dashboard panel",
        "1. Open performance graphs tab",
        "N/A",
        "Graphs load, showing CPU utilization, database sessions counts, and API response delay metrics.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Admin Dashboard", "Feedback Management", "Verify display of feedback list submitted by users",
        "Admin is on the Feedback panel page",
        "1. Open Feedback table logs list",
        "N/A",
        "Feedback items display, listing User ID, Ratings scores (1-5), Comments, and Timestamp.",
        "High", "Major", "Functional", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Admin Dashboard", "Feedback Management", "Verify filtering feedback list by rating",
        "Admin is on the Feedback panel page",
        "1. Click Filter by Rating dropdown\n2. Select '1 Star'\n3. Select '5 Stars'",
        "N/A",
        "Feedback table updates to list only records matching selected star ratings.",
        "Medium", "Minor", "UI", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "Feedback Management", "Verify feedback status update toggle",
        "Admin is reviewing feedback list",
        "1. Click 'Mark as Resolved' button on one feedback row",
        "Feedback Row ID: 4",
        "Status icon changes to green checkmark. Record updates status to 'Resolved' in database.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "Feedback Management", "Verify deletion of offensive feedback comments",
        "Feedback list displays user feedback entry with inappropriate text",
        "1. Click the 'Delete' trashbin icon on the target row\n2. Confirm delete popup",
        "Feedback ID: 8",
        "Row is removed from list view and deleted from database record.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify exporting user lists to CSV file format",
        "Admin is viewing the Users List screen",
        "1. Click 'Export to CSV' button",
        "N/A",
        "Browser/Client initiates file download of 'user_list.csv' containing full table records data.",
        "Medium", "Minor", "Functional", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "Analytics", "Verify Admin Dashboard layouts adapt to tablet screens size",
        "Admin dashboard is running on Tablet view size",
        "1. Inspect side panel folding states and chart layouts sizing",
        "Screen size: 10 inches",
        "Side navigation menu folds to icons bar, cards stretch to fill space evenly.",
        "Low", "Minor", "UI", "No", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify edit system settings values",
        "Admin is logged in and opens settings menu",
        "1. Change default system message banner text\n2. Save settings",
        "Settings text update",
        "Settings update successfully. Changes display on standard user home page.",
        "Medium", "Major", "Functional", "Yes", ["Regression"]
    ),
    (
        "Admin Dashboard", "Admin Login", "Verify Admin idle session automatic logout locks",
        "Admin dashboard page remains active but idle for 30 minutes",
        "1. Leave dashboard open without user cursor movement/clicks",
        "N/A",
        "Admin session terminates, page locks. User is redirected to Login view.",
        "High", "Major", "Security", "No", ["Regression"]
    ),
    (
        "Admin Dashboard", "Admin Login", "Verify Admin account login with Multi-Factor Authentication",
        "MFA is activated for Admin profile login keys",
        "1. Enter email/password\n2. Press login\n3. Enter valid MFA code sent to admin device",
        "MFA Code: 554433",
        "Admin dashboard opens successfully. Blocks dashboard access for incorrect code.",
        "High", "Critical", "Security", "No", ["Regression"]
    ),
    (
        "Admin Dashboard", "User Management", "Verify admin audit logs logs registration checks",
        "Admin disabled user Ramesh account",
        "1. Open admin audit logs database table",
        "N/A",
        "Audit row records action: 'Admin disabled user account ID: 15' with correct timestamp.",
        "High", "Major", "Database", "Yes", ["Regression"]
    ),

    # L. Performance Testing (25 Test Cases)
    (
        "Performance", "Scale Upload", "Verify API response latency during spike user registrations",
        "Stress testing script configured",
        "1. Send 50 concurrent login API calls in 1 second\n2. Log individual response duration",
        "N/A",
        "Average login API response duration stays under 500ms. No connection timeouts.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Scale Upload", "Verify upload queue stability under concurrent batch uploads",
        "Multiple tabs active",
        "1. Upload 10 image files simultaneously from 10 client sessions",
        "10 files (2MB each)",
        "FastAPI processes all files without memory leaks, allocating paths correctly.",
        "Medium", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Scale OCR", "Verify standard document OCR duration times",
        "User is on guidance window",
        "1. Click process on 1-page JPG file\n2. Stop timer when text boxes render",
        "Standard PAN JPG form",
        "Processing text extraction completes within 5.0 seconds under standard bandwidth.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Scale OCR", "Verify multi-page PDF document OCR processing duration limit",
        "Large document uploaded",
        "1. Run OCR processing on a 5-page PDF document file",
        "5-page PDF (4.5MB)",
        "EasyOCR completes parsing within 15 seconds, loading pages sequence details correctly.",
        "Medium", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Scale OCR", "Verify EasyOCR engine worker RAM memory leak checks",
        "FastAPI OCR script running",
        "1. Run 100 consecutive document OCR requests on FastAPI server\n2. Check server memory log data",
        "N/A",
        "Memory usage rises but falls after GC sweep, confirming zero lingering memory leaks.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify FastAPI host CPU utilization peak levels",
        "Server receives constant load requests",
        "1. Run load generator simulation at 10 requests per second for 10 minutes",
        "N/A",
        "CPU utilization stays below 80% on Render host nodes.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Database Performance", "Verify database query response speed on large record datasets",
        "Users table filled with 100,000 mock rows",
        "1. Query User listing details API\n2. Measure database query execution times",
        "N/A",
        "Select queries return data under 100ms using index keys.",
        "Medium", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Scale TTS", "Verify TTS (Text-to-Speech) file generation duration response",
        "TTS api active",
        "1. Trigger TTS generation for 500 characters of guidance text",
        "500 characters text",
        "Audio byte stream starts within 1.5 seconds from FastAPI endpoint.",
        "Medium", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify network bandwidth usage for audio playback streams",
        "Audio streaming active in-app",
        "1. Measure network bytes usage during playing 1-minute audio guide",
        "N/A",
        "Streams lightweight compressed audio formats, consuming less than 1MB bandwidth per minute.",
        "Low", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify page load speeds for Admin analytics dashboards page",
        "Admin is logged in",
        "1. Navigate to Analytics dashboard\n2. Stop timer on complete chart rendering",
        "N/A",
        "Charts and dashboard load completed within 2.0 seconds.",
        "Medium", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Database Performance", "Verify database connection pool timeout recovery limits",
        "Connection pool limit set to 20",
        "1. Run 25 long-running query transactions concurrently",
        "N/A",
        "5 queue overflow requests wait or return timeout error gracefully within 10s.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify Render host auto-scaling triggers metrics values",
        "Render instances configured with auto-scale rules",
        "1. Flood server with high volume traffic (200 requests/sec)",
        "N/A",
        "CPU load peaks past 75%, and Render spins up additional instance nodes to balance traffic load.",
        "Low", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify android device battery consumption metrics",
        "Kotlin app active on mobile device",
        "1. Run voice guide wizard continuously for 30 minutes\n2. Read battery drainage logs",
        "N/A",
        "Battery consumption is less than 3% over 30 minutes of continuous voice guide play.",
        "Low", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify android app RAM allocation spikes limits",
        "App active",
        "1. Navigate: upload -> OCR -> zoom guidance -> home list\n2. Monitor memory allocation",
        "N/A",
        "Android app memory allocation remains stable, verifying zero context leaks.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Database Performance", "Verify database performance during parallel schema update scripts",
        "Alembic migration runs on staging DB",
        "1. Run write queries in-app while migration executes",
        "N/A",
        "Database processes read/write queries successfully; locking is minimized.",
        "Low", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify fastapi server cold start initialization speeds",
        "Render instance inactive for 30 minutes (scales to zero)",
        "1. Click login submit\n2. Measure response delay",
        "N/A",
        "First call launches instance, completing request in under 15 seconds (documented cold start limit).",
        "Medium", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify CDN asset cache hit ratios percentages",
        "Static forms and images stored on CDN",
        "1. Send 10 requests to download static form images templates\n2. Read response header metrics",
        "N/A",
        "Headers display X-Cache: HIT for subsequent queries, returning files instantly.",
        "Low", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Database Performance", "Verify redis cache hit ratios performance speed checks",
        "FastAPI caching active",
        "1. Read User Profile details twice in sequence\n2. Log DB query count metrics",
        "N/A",
        "Second request reads from Redis cache in under 5ms, avoiding database roundtrip.",
        "Medium", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify file upload stability on slow network connections",
        "Client network speed limited to 3G slow (300Kbps)",
        "1. Upload 3MB document image file",
        "File: 3MB JPG",
        "Upload takes time but completes successfully without network connection breaks.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify Android app launch cold start time limits",
        "App completely closed",
        "1. Tap App icon from launcher\n2. Stop timer on complete login view display",
        "N/A",
        "Login screen is completely rendered and interactive within 1.8 seconds.",
        "Medium", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Scale OCR", "Verify queue scaling with multiple backend EasyOCR processing threads",
        "10 parallel OCR requests sent simultaneously",
        "1. Monitor OCR execution list queue",
        "N/A",
        "Requests are processed sequentially or via multithreading without task dropouts.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "Database Performance", "Verify database write latency times for feedback entries logs",
        "100 users submit feedback concurrently",
        "1. Measure database execution delay",
        "N/A",
        "Average feedback database save transaction latency remains under 80ms.",
        "Medium", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify viewport zoom FPS rendering rates metrics",
        "Guidance page open, document image loaded",
        "1. Pinch-to-zoom in and out rapidly\n2. Read GPU rendering frame drops profile",
        "N/A",
        "Rendering frame rate stays above 55 FPS. Zero layout stuttering or screen freezes.",
        "Low", "Minor", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify long running soak test stability logs checks",
        "Server running continuously for 24 hours under low-intensity background load",
        "1. Inspect logs and memory profile stats after 24 hours run",
        "N/A",
        "FastAPI instance remains running. Active database connections count remains stable.",
        "High", "Major", "Performance", "No", ["Regression"]
    ),
    (
        "Performance", "System Performance", "Verify download file size of additional multi-language voice packs",
        "User toggles Tamil voice guidance",
        "1. MonitorTamil voice data pack payload bytes download",
        "N/A",
        "Tamil localized dictionary JSON pack is under 150KB size, conserving storage.",
        "Low", "Minor", "Performance", "No", ["Regression"]
    ),

    # M. Regression Testing (30 Test Cases)
    (
        "Regression", "Core Journeys", "Verify full user journey from signup to dashboard view",
        "User is not registered",
        "1. Sign up user\n2. Select language English\n3. View dashboard screens",
        "Ramesh, ramesh.g@example.com, Password@123",
        "Signup completes, login succeeds automatically, English language is pre-selected on home dashboard.",
        "High", "Critical", "Regression", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Regression", "Core Journeys", "Verify profile details editing persistence across restarts",
        "User has updated user details",
        "1. Modify name to Ramesh Goud\n2. Close app\n3. Launch app & verify",
        "New Name: Ramesh Goud",
        "Updated name details persist in database and display accurately in profile screen.",
        "High", "Critical", "Regression", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Regression", "Core Journeys", "Verify document upload and OCR guidance flow",
        "User is logged in",
        "1. Upload SBI bank form\n2. Execute OCR text extraction\n3. Navigate steps guidance",
        "File: sbi.jpg",
        "Document uploaded, OCR outputs structured text fields, guidance highlights overlays perfectly.",
        "High", "Critical", "Regression", "Yes", ["Smoke", "Sanity", "Regression", "Critical Path"]
    ),
    (
        "Regression", "Core Journeys", "Verify voice guidance playback controls switch languages",
        "User is on document guidance page",
        "1. Start playing voice guide\n2. Switch language from English to Hindi mid-way",
        "N/A",
        "English voice stops instantly; Hindi translation streams and begins playing from same index.",
        "High", "Critical", "Regression", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Regression", "Admin Operations", "Verify admin search and user status disabling flow",
        "Admin is logged into panel",
        "1. Search 'ramesh_g'\n2. Click Ramesh row\n3. Select Disable Account\n4. Save changes",
        "User: ramesh_g",
        "User account status changes to Blocked in DB. User cannot log in on device A.",
        "High", "Critical", "Regression", "Yes", ["Sanity", "Regression", "Critical Path"]
    ),
    (
        "Regression", "Database Migrations", "Verify login logic functions correctly post database migrations",
        "Alembic DB upgrades completed on server",
        "1. Launch application\n2. Submit valid user login details",
        "N/A",
        "User logs in successfully. Migration preserves table schemas and user credential hashes.",
        "High", "Critical", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "API Caching", "Verify API caching policies do not serve stale user data",
        "User profile updated",
        "1. Open profile page\n2. Change full name\n3. Navigate to Home and then back to profile",
        "N/A",
        "Updated profile details display immediately. Cache headers instruct client to fetch fresh copy.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "UI Themes", "Verify screen theme preferences persist across logins",
        "User has set app layout to Dark Mode",
        "1. Logout of app\n2. Login with valid user credentials\n3. Inspect app theme",
        "N/A",
        "Dark mode settings remain active. Layout renders in dark theme by default.",
        "Medium", "Minor", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "UI Navigation", "Verify back button stack depth limits",
        "User navigating screens: Dashboard -> Upload -> Guide -> Home",
        "1. Press hardware back button three times in sequence",
        "N/A",
        "Returns back through views step-by-step. Does not cause app force-close errors.",
        "Medium", "Minor", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "Voice Guidance", "Verify Telugu voice guide playback stability",
        "Telugu language selected",
        "1. Start playing Telugu voice guide on bank form",
        "Telugu language selection",
        "Audio plays with correct pronunciation, matching text overlay perfectly.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Voice Guidance", "Verify Hindi voice guide playback stability",
        "Hindi language selected",
        "1. Start playing Hindi voice guide on PAN form",
        "Hindi language selection",
        "Audio plays clearly without interruptions.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Voice Guidance", "Verify Tamil voice guide playback stability",
        "Tamil language selected",
        "1. Start playing Tamil voice guide on Aadhar form",
        "Tamil language selection",
        "Audio plays clearly without interruptions.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Voice Guidance", "Verify English voice guide playback stability",
        "English language selected",
        "1. Start playing English voice guide on PAN form",
        "English language selection",
        "Audio plays clearly without interruptions.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Security Policies", "Verify api request header validation rules",
        "Server libraries upgraded",
        "1. Call POST `/api/documents/upload` containing security headers",
        "Authorization headers",
        "Requests process successfully. Updates did not introduce CORS or validation failures.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Admin Operations", "Verify admin analytics chart layouts display",
        "Charts libraries updated to new versions",
        "1. Open Admin Analytics dashboard page",
        "N/A",
        "Charts and dashboard load cleanly. Registration graphs render without errors.",
        "Medium", "Minor", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "Feedback System", "Verify user rating submission calculations",
        "Multiple ratings submitted",
        "1. Submit feedback ratings on multiple sessions\n2. Open Admin statistics",
        "N/A",
        "Average score increments correctly in feedback summary reports.",
        "Medium", "Minor", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Form Upload", "Verify multi-page PDF document parsing consistency",
        "PDF parser libraries updated",
        "1. Upload a 3-page bank PDF document",
        "N/A",
        "Document uploads and parses into 3 separate guidance pages correctly.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Offline Sync", "Verify app behavior on sudden offline transition",
        "User is inside OCR processing screen",
        "1. Terminate network connection mid-process\n2. Check feedback display",
        "N/A",
        "Spinner terminates. Screen alerts: 'Network lost. Check settings.' User can re-upload when back online.",
        "High", "Major", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "UI Navigation", "Verify deep links routing constraints",
        "Application is closed on phone",
        "1. Tap a deep link URL pointing to profile settings screen",
        "Deep link: formsahayak://settings",
        "App starts, prompts login if session is empty, and navigates directly to Settings screen.",
        "Medium", "Minor", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "Database Integrity", "Verify foreign key constraint cascades",
        "User account deleted by admin",
        "1. Admin deletes user account\n2. Check documents and feedback DB rows",
        "User ID: 15",
        "Database is clean. Zero orphaned records left in DB table rows.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Form History", "Verify deletion of documents from Form History logs",
        "User is on Form History screen",
        "1. Slide document item left or tap trash icon\n2. Confirm delete",
        "Document ID: 15",
        "Item is dropped from UI list immediately. Row deleted from SQLite/Cloud database records.",
        "High", "Major", "Regression", "Yes", ["Sanity", "Regression"]
    ),
    (
        "Regression", "UI Rendering", "Verify Jetpack Compose rendering stability under high activity",
        "User is browsing rapid updates",
        "1. Open details view and scroll up and down page rapidly",
        "N/A",
        "App does not freeze or trigger Compose OutOfMemory exceptions.",
        "Medium", "Minor", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "Database Integrity", "Verify SSL connection configurations to MySQL Cloud",
        "Database migration executed",
        "1. Verify database driver settings",
        "N/A",
        "Connections to Railway/Aiven MySQL enforce SSL, rejecting plain text communication.",
        "High", "Critical", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "Authentication", "Verify session token validation check frequency",
        "User is active inside guidance wizard",
        "1. Attempt accessing backend API after session token was revoked manually in database",
        "N/A",
        "Next API call is blocked, displaying: 'Session expired. Please log in again.'",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Offline Sync", "Verify local offline cache database validation",
        "User has offline access to some standard templates",
        "1. Open app while offline\n2. Try rendering template help text",
        "N/A",
        "Cached guidance loads from local SQLite/Room database cleanly.",
        "Medium", "Minor", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "UI Themes", "Verify dynamic layouts on high resolution screens",
        "App updated",
        "1. View about developer details on 12-inch tablet screen",
        "N/A",
        "Layout remains clean. Text fields are sized correctly.",
        "Low", "Minor", "Regression", "No", ["Regression"]
    ),
    (
        "Regression", "Security Policies", "Verify clickjacking headers validation metrics",
        "Web pages code modified",
        "1. Load Admin dashboard inside iframe on rogue site",
        "N/A",
        "Browser blocks loading, showing error about X-Frame-Options or Content-Security-Policy.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Database Integrity", "Verify rollback of transactions during partial API failures",
        "API fails during multiple table records updates",
        "1. Run profile image update with database session mock failures",
        "N/A",
        "Database rolls back, restoring old profile details consistently.",
        "High", "Major", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "API Caching", "Verify cache control response header configurations",
        "API updates completed",
        "1. Call static forms metadata list endpoint\n2. Inspect response headers",
        "N/A",
        "Headers contain Cache-Control: max-age=3600 public for static templates lists.",
        "Medium", "Minor", "Regression", "Yes", ["Regression"]
    ),
    (
        "Regression", "Core Journeys", "Verify graceful handling of out-of-memory simulations",
        "High processing load active",
        "1. Run memory intensive operations while app is in background",
        "N/A",
        "Android OS reclaims memory. Re-opening app restores state or starts cleanly without crashing.",
        "Medium", "Major", "Regression", "No", ["Regression"]
    )
]

def style_worksheet(ws):
    # Set Gridlines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Freeze the top header row
    ws.freeze_panes = "A2"
    
    # Define color fills (using professional colors)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark steel blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Very light gray
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    header_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )
    
    # Priority colors
    priority_styles = {
        "High": {"fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="C00000", bold=True)},
        "Medium": {"fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="7F6000", bold=True)},
        "Low": {"fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="375623", bold=True)}
    }
    
    # Severity colors
    severity_styles = {
        "Blocker": {"fill": PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="900C3F", bold=True)},
        "Critical": {"fill": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="C00000", bold=True)},
        "Major": {"fill": PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="B25900", bold=True)},
        "Minor": {"fill": PatternFill(start_color="D1F2EB", end_color="D1F2EB", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="117A65", bold=True)}
    }
    
    # Automation colors
    auto_styles = {
        "Yes": {"fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="375623", bold=True)},
        "No": {"fill": PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="5D6D7E")}
    }

    # Style Header row
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, 13):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    # Style Data rows
    num_rows = ws.max_row
    for r in range(2, num_rows + 1):
        ws.row_dimensions[r].height = 24
        is_even = (r % 2 == 0)
        row_fill = zebra_fill if is_even else white_fill
        
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=10)
            
            # Default text alignment
            if c in [1, 9, 10, 11, 12]:  # ID, Priority, Severity, Type, Auto
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [2, 3]:  # Module, Feature
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = row_fill
            else:  # Text heavy columns
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = row_fill
                
            # Apply color highlights
            if c == 9:  # Priority
                val = cell.value
                if val in priority_styles:
                    cell.fill = priority_styles[val]["fill"]
                    cell.font = priority_styles[val]["font"]
            elif c == 10:  # Severity
                val = cell.value
                if val in severity_styles:
                    cell.fill = severity_styles[val]["fill"]
                    cell.font = severity_styles[val]["font"]
            elif c == 12:  # Automation Candidate
                val = cell.value
                if val in auto_styles:
                    cell.fill = auto_styles[val]["fill"]
                    cell.font = auto_styles[val]["font"]
            elif c == 1: # ID
                cell.font = Font(name="Segoe UI", size=10, bold=True)
                cell.fill = row_fill

    # Set column width constraints
    # Keep some columns auto-fit, and wrap text-heavy columns
    col_widths = {
        1: 15,  # Test Case ID
        2: 20,  # Module
        3: 22,  # Feature
        4: 35,  # Test Scenario
        5: 35,  # Precondition
        6: 45,  # Test Steps
        7: 35,  # Test Data
        8: 45,  # Expected Result
        9: 12,  # Priority
        10: 12, # Severity
        11: 15, # Test Type
        12: 20  # Automation Candidate
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Enable autofilter
    ws.auto_filter.ref = f"A1:L{num_rows}"

def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Helper to populate test cases in a sheet
    def populate_sheet(title, filter_fn=None):
        ws = wb.create_sheet(title=title)
        
        # Write Headers
        headers = [
            "Test Case ID", "Module", "Feature", "Test Scenario", 
            "Precondition", "Test Steps", "Test Data", "Expected Result", 
            "Priority", "Severity", "Test Type", "Automation Candidate (Yes/No)"
        ]
        ws.append(headers)
        
        # Write filtered test cases
        row_counter = 1
        
        # Module code mapping for ID generation
        module_codes = {
            "Authentication": "AUTH",
            "Profile": "PROF",
            "OCR": "OCR",
            "Form Upload": "FORM",
            "Guidance": "GUID",
            "Voice Guidance": "VOIC",
            "Database": "DB",
            "API": "API",
            "UI": "UI",
            "Security": "SEC",
            "Admin Dashboard": "ADM",
            "Performance": "PERF",
            "Regression": "REG"
        }
        
        # Track sequence count per module code to generate unique IDs
        seq_counters = {code: 0 for code in module_codes.values()}
        
        for tc in tc_list:
            module, feature, scenario, precondition, steps, data, expected, priority, severity, ttype, auto, suites = tc
            
            # Check if this test case passes the suite filter
            if filter_fn and not filter_fn(suites):
                continue
                
            code = module_codes.get(module, "GEN")
            seq_counters[code] += 1
            tc_id = f"TC-{code}-{seq_counters[code]:03d}"
            
            row_data = [
                tc_id, module, feature, scenario, precondition, 
                steps, data, expected, priority, severity, ttype, auto
            ]
            ws.append(row_data)
            
        style_worksheet(ws)
        print(f"Generated Sheet: {title} with {ws.max_row - 1} test cases.")

    # 1. All Test Cases Sheet (All 335 items)
    populate_sheet("All Test Cases", filter_fn=None)
    
    # 2. Smoke Test Suite Sheet
    populate_sheet("Smoke Test Suite", filter_fn=lambda suites: "Smoke" in suites)
    
    # 3. Sanity Test Suite Sheet
    populate_sheet("Sanity Test Suite", filter_fn=lambda suites: "Sanity" in suites)
    
    # 4. Regression Test Suite Sheet (all test cases)
    populate_sheet("Regression Test Suite", filter_fn=lambda suites: "Regression" in suites)
    
    # 5. Critical Path Test Suite Sheet
    populate_sheet("Critical Path Test Suite", filter_fn=lambda suites: "Critical Path" in suites)
    
    # Save Workbook
    filename = "FormSahayak_Test_Cases.xlsx"
    wb.save(filename)
    print(f"Successfully generated styled Excel workbook: {filename}")

if __name__ == "__main__":
    generate_workbook()
